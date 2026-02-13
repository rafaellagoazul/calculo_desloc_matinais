import os
import sys
import subprocess
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


import customtkinter as ctk
COLUNAS = [
    "SV",
    "DIA",
    "COD VENDEDOR",
    "COD CLIENTE",
    "CASA VEND",
    "DISTRIBUIDOR",
    "1ºCLIENTE",
    "CASA → DIST",
    "CASA → CLI",
    "DIFERENÇA",
]



def _abrir_arquivo(path: str):
    if sys.platform.startswith("win"):
        os.startfile(path)
    elif sys.platform.startswith("darwin"):
        subprocess.call(["open", path])
    else:
        subprocess.call(["xdg-open", path])


def baixar_modelo_excel(parent):
    path = filedialog.asksaveasfilename(
        parent=parent,
        title="Salvar modelo Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="MODELO_DESLOCAMENTOS.xlsx"
    )

    if not path:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Deslocamentos"

    ws.append(COLUNAS)

    bold = Font(bold=True)
    normal = PatternFill("solid", fgColor="DDDDDD")
    calculado = PatternFill("solid", fgColor="EEEEEE")

    for i, nome in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = bold
        cell.fill = calculado if nome in {
            "CASA → DIST", "CASA → CLI", "DIFERENÇA"
        } else normal
        ws.column_dimensions[cell.column_letter].width = 18

    try:
        wb.save(path)
    except Exception:
        _modal_erro_arquivo_em_uso(parent)
        return

    # 👇 AGORA SIM
    _modal_confirmar_abertura(parent, path)


def _modal_confirmar_abertura(parent, path: str):
    modal = ctk.CTkToplevel(parent)
    modal.title("Modelo criado")
    modal.geometry("420x200")
    modal.resizable(False, False)

    modal.transient(parent)
    modal.grab_set()

    modal.update_idletasks()
    x = parent.winfo_x() + parent.winfo_width() // 2 - 210
    y = parent.winfo_y() + parent.winfo_height() // 2 - 100
    modal.geometry(f"+{x}+{y}")

    frame = ctk.CTkFrame(modal)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    ctk.CTkLabel(
        frame,
        text="Modelo Excel criado com sucesso.\n\nDeseja abrir o arquivo agora?",
        justify="center"
    ).pack(pady=20)

    botoes = ctk.CTkFrame(frame)
    botoes.pack(pady=10)

    ctk.CTkButton(
        botoes,
        text="📂 Abrir",
        command=lambda: (_abrir_arquivo(path), modal.destroy())
    ).pack(side="left", padx=10)

    ctk.CTkButton(
        botoes,
        text="Fechar",
        fg_color="#555",
        command=modal.destroy
    ).pack(side="left", padx=10)


def _modal_erro_arquivo_em_uso(parent):
    modal = ctk.CTkToplevel(parent)
    modal.title("Arquivo em uso")
    modal.geometry("420x180")
    modal.resizable(False, False)

    modal.transient(parent)
    modal.grab_set()

    modal.update_idletasks()
    x = parent.winfo_x() + parent.winfo_width() // 2 - 210
    y = parent.winfo_y() + parent.winfo_height() // 2 - 90
    modal.geometry(f"+{x}+{y}")

    frame = ctk.CTkFrame(modal)
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    ctk.CTkLabel(
        frame,
        text=(
            "Não foi possível salvar o arquivo.\n\n"
            "O modelo está aberto ou sem permissão de escrita.\n"
            "Feche o arquivo e tente novamente."
        ),
        justify="center"
    ).pack(pady=20)

    ctk.CTkButton(
        frame,
        text="OK",
        width=120,
        command=modal.destroy
    ).pack(pady=10)
